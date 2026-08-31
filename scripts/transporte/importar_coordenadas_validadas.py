#!/usr/bin/env python3
"""Importa únicamente coordenadas marcadas como "Ubicada" en la planilla de rastreo.

El cruce se hace por el ID estable de la parada. No aplica semejanza de nombres ni
resuelve homónimos automáticamente. También conserva un catálogo JSON auditable.
"""
from __future__ import annotations

import argparse
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


def norm(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    return re.sub(r"[^A-Z0-9]+", " ", "".join(c for c in text if unicodedata.category(c) != "Mn").upper()).strip()


def read_verified(workbook: Path) -> list[dict]:
    book = load_workbook(workbook, read_only=True, data_only=True)
    sheet = book["Paradas pendientes"]
    headers = {str(cell.value).strip(): index for index, cell in enumerate(sheet[3], 1) if cell.value}
    required = ["Estado", "Localidad / parada", "ID", "Corredores", "Latitud", "Longitud", "Fuente de verificación"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError("Faltan columnas requeridas: " + ", ".join(missing))
    records = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        def value(name):
            return row[headers[name] - 1]
        if norm(value("Estado")) != "UBICADA":
            continue
        lat, lon = value("Latitud"), value("Longitud")
        if not isinstance(lat, (int, float)) or not isinstance(lon, (int, float)):
            raise ValueError(f"Coordenadas inválidas para {value('Localidad / parada')}")
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            raise ValueError(f"Coordenadas fuera de rango para {value('Localidad / parada')}")
        records.append({
            "id": str(value("ID")).strip(),
            "name": str(value("Localidad / parada")).strip(),
            "corridors": [item.strip() for item in str(value("Corredores") or "").split("|") if item.strip()],
            "lat": float(lat),
            "lon": float(lon),
            "source": str(value("Fuente de verificación") or "Planilla de rastreo validada").strip(),
        })
    return records


def apply(records: list[dict], routes: dict, locations: dict) -> tuple[int, int]:
    route_count = 0
    location_count = 0
    for record in records:
        place = routes.get("places", {}).get(record["id"])
        if place:
            place.update(lat=record["lat"], lon=record["lon"], geo_source=record["source"])
            route_count += 1
        if record["id"].startswith("pdf-"):
            key = next((name for name in locations.get("unresolved", []) if norm(name) == norm(record["name"])), record["name"].upper())
            locations.setdefault("locations", {})[key] = {
                "lat": record["lat"], "lon": record["lon"], "source": record["source"], "source_name": record["name"]
            }
            locations["unresolved"] = [name for name in locations.get("unresolved", []) if norm(name) != norm(record["name"])]
            location_count += 1
    linked_ids = set(routes.get("bindings", {}).values())
    linked_places = {stop["place_id"] for profile in routes.get("profiles", []) if profile["id"] in linked_ids for stop in profile["stops"]}
    stats = routes.setdefault("stats", {})
    stats["geolocated_places"] = sum(place.get("lat") is not None and place.get("lon") is not None for place in routes.get("places", {}).values())
    stats["linked_places_unlocated"] = sum(routes["places"][pid].get("lat") is None for pid in linked_places)
    routes["validated_coordinates"] = {
        "count": len(records),
        "applied_to_routes": route_count,
        "applied_to_pdf_endpoints": location_count,
        "policy": "Cruce por ID estable; sin coincidencias difusas ni unión automática de homónimos",
    }
    return route_count, location_count


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--routes", type=Path, required=True)
    parser.add_argument("--locations", type=Path, required=True)
    parser.add_argument("--catalog", type=Path, required=True)
    args = parser.parse_args()
    records = read_verified(args.workbook)
    routes = json.loads(args.routes.read_text(encoding="utf-8"))
    locations = json.loads(args.locations.read_text(encoding="utf-8"))
    route_count, location_count = apply(records, routes, locations)
    args.routes.write_text(json.dumps(routes, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    args.locations.write_text(json.dumps(locations, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    args.catalog.parent.mkdir(parents=True, exist_ok=True)
    args.catalog.write_text(json.dumps({
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "policy": "Solo registros marcados como Ubicada; cruce por ID estable",
        "points": records,
    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"verified": len(records), "routes": route_count, "pdf_endpoints": location_count}, ensure_ascii=False))


if __name__ == "__main__":
    main()
